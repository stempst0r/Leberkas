import tkinter as tk
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import sqlite3
import sys
import os


class GUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Leberkas")
        self.root.geometry("800x400")  # Set the initial size of the window
        self.root.resizable(0, 0) # Fixate the size of the window

        self.selected_files = set()  # Maintain a set of selected files

        # Create a frame for file selection and display
        file_frame = tk.Frame(self.root)
        file_frame.pack(fill=tk.BOTH, expand=True)

        # Listbox to display selected filenames (not full paths)
        self.file_listbox = tk.Listbox(file_frame)
        self.file_listbox.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)

        # Scrollbar for the file_listbox on the right side
        scrollbar = tk.Scrollbar(file_frame, orient=tk.VERTICAL)
        scrollbar.config(command=self.file_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.file_listbox.config(yscrollcommand=scrollbar.set)

        # Create a frame for buttons
        button_frame = tk.Frame(self.root)
        button_frame.pack(fill=tk.X)

        # Button to add CSV files
        self.add_button = tk.Button(button_frame, text="Add CSV", command=self.select_files)
        self.add_button.pack(side=tk.LEFT)

        # Button to remove selected CSV files
        self.remove_button = tk.Button(button_frame, text="Remove CSV", command=self.delete_selected)
        self.remove_button.pack(side=tk.LEFT)

        # Button to import CSV files
        self.import_button = tk.Button(button_frame, text="Import", command=self.import_files)
        self.import_button.pack(side=tk.LEFT)

        # Button to Generate XLSX
        self.generate_button = tk.Button(button_frame, text="Generate XLSX", command=self.generate_xlsx)
        self.generate_button.pack(side=tk.RIGHT)

        # Create a frame for terminal output
        terminal_frame = tk.Frame(self.root)
        terminal_frame.pack(fill=tk.BOTH, expand=True)

        # Text widget to display terminal output with a scrollbar
        self.terminal_text = tk.Text(terminal_frame, wrap=tk.WORD, bg="black", fg="lime")
        self.terminal_text.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)

        # Configure a tag for the green text
        self.terminal_text.tag_configure("lime", foreground="lime")

        # Scrollbar for the terminal_text on the right side
        terminal_scrollbar = tk.Scrollbar(terminal_frame, orient=tk.VERTICAL)
        terminal_scrollbar.config(command=self.terminal_text.yview)
        terminal_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.terminal_text.config(yscrollcommand=terminal_scrollbar.set)

        # Redirect stdout to update the Text widget
        sys.stdout = self
    
    def on_closing(self):
        # Restore sys.stdout to its original value before closing the window
        sys.stdout = sys.__stdout__

        # Delete the SQLite database file if it exists
        if os.path.exists("data.db"):
            os.remove("data.db")

        # Close the Tkinter window
        self.root.destroy()

    def select_files(self):
        files = filedialog.askopenfilenames(filetypes=[("CSV files", "*.csv")])
        for file in files:
            if file not in self.selected_files:
                self.selected_files.add(file)
                # Add only the filename (not the full path) to the listbox
                filename = file.split("/")[-1]  # Extract the filename from the path
                self.file_listbox.insert(tk.END, filename)

    def delete_selected(self):
        selected_indices = self.file_listbox.curselection()
        for index in selected_indices:
            filename = self.file_listbox.get(index)
            file = next((f for f in self.selected_files if filename in f), None)
            if file:
                self.selected_files.remove(file)  # Remove the file from the set
                self.file_listbox.delete(index)

    def import_files(self):
        for file in self.selected_files.copy():  # Create a copy to avoid modifying while iterating
            import_csv(file)  # Call the import_csv function from data_import.py
            self.selected_files.remove(file)  # Remove the file from the selected set

    def generate_xlsx(self):
        # Call the export_xlsx function from data_export.py
        export_xlsx("data.db", "export")
        self.write("XLSX files generated successfully.\n")

    def write(self, text):
        # This method is called when sys.stdout.write() is used
        self.terminal_text.insert(tk.END, text, "lime")  # Apply "green" tag to the text
        self.terminal_text.see(tk.END)  # Scroll to the end of the text

def create_database_if_not_exists():
    # Check if the database file exists
    if not os.path.exists("data.db"):
        # If it doesn't exist, create a new database
        conn = sqlite3.connect("data.db")
        conn.close()
        print("Created a new database: data.db")

def create_table_if_not_exists():
    conn = sqlite3.connect("data.db")
    cursor = conn.cursor()

    # Check if the 'imported_data' table exists
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='imported_data'")
    table_exists = cursor.fetchone()

    if not table_exists:
        # If it doesn't exist, create the 'imported_data' table
        cursor.execute('''
            CREATE TABLE imported_data (
                id INTEGER PRIMARY KEY,
                date DATETIME,
                asset TEXT,
                txid TEXT,
                party_a TEXT,
                party_b TEXT,
                address_a TEXT,
                address_b TEXT,
                amount_a NUMERIC,
                amount_b NUMERIC,
                usd_a NUMERIC,
                usd_b NUMERIC
            )
        ''')
        conn.commit()
        print("Created the 'imported_data' table.")

    conn.close()

def import_csv(file_path):
    create_database_if_not_exists()
    create_table_if_not_exists()

    # Connect to the SQLite database
    conn = sqlite3.connect("data.db")
    cursor = conn.cursor()

    try:
        with open(file_path, "r", encoding="utf-8") as csv_file:
            lines = csv_file.readlines()

            # Extract party A and party B from lines 5 and 10, Column 1
            party_a = lines[4].strip().split(",")[0]
            party_b = lines[9].strip().split(",")[0]

            # Initialize a list to store data rows
            data_rows = []

            # Iterate from line 28 until an empty line is encountered
            i = 27  # Start from line 28 (0-based index)
            while i < len(lines):
                line = lines[i].strip()
                if not line:  # Stop when an empty line is encountered
                    break
                parts = line.split(",")

                # Extract values from the CSV
                asset = parts[0]
                txid = parts[1]
                date = parts[2]
                address_a = parts[3]
                address_b = parts[4]
                amount_a = parts[5]
                amount_b = parts[6]
                usd_a = parts[7]
                usd_b = parts[8]

                # Check if a row with the same values (except the primary key) exists
                cursor.execute('''
                    SELECT * FROM imported_data
                    WHERE
                        date = ? AND asset = ? AND txid = ? AND party_a = ? AND party_b = ?
                        AND address_a = ? AND address_b = ? AND amount_a = ? AND amount_b = ?
                        AND usd_a = ? AND usd_b = ?
                ''', (date, asset, txid, party_a, party_b, address_a, address_b,
                      amount_a, amount_b, usd_a, usd_b))

                existing_row = cursor.fetchone()
                if existing_row is not None:
                    print(f"Row with duplicate values found (Primary Key: {existing_row[0]}). Skipping import.")
                else:
                    # Append the extracted data to the data_rows list
                    data_rows.append((
                        date, asset, txid, party_a, party_b, address_a, address_b,
                        amount_a, amount_b, usd_a, usd_b
                    ))

                i += 1  # Move to the next line

            # Insert the data into the SQLite database
            cursor.executemany('''
                INSERT OR IGNORE INTO imported_data
                (date, asset, txid, party_a, party_b, address_a, address_b,
                amount_a, amount_b, usd_a, usd_b)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', data_rows)
            conn.commit()

            print(f"Imported data from {file_path} into the database.")

    except Exception as e:
        print(f"Error importing data from {file_path}: {str(e)}")

    finally:
        conn.close()

def get_service_providers(db_file):
    # Connect to the SQLite database
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()

    # Fetch unique service providers from both party_a and party_b columns using UNION
    cursor.execute("SELECT party_a FROM imported_data WHERE party_a LIKE '%.%' UNION SELECT party_b FROM imported_data WHERE party_b LIKE '%.%'")
    results = cursor.fetchall()

    # Collect the unique service providers into a set
    service_providers = set(result[0] for result in results)

    # Close the database connection
    conn.close()

    return service_providers

def fetch_rows_by_service_provider(db_file, service_provider):
    # Connect to the SQLite database
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()

    # Fetch rows where the service provider is found in party_a or party_b
    cursor.execute("SELECT * FROM imported_data WHERE party_a LIKE ? OR party_b LIKE ?", ('%' + service_provider + '%', '%' + service_provider + '%'))
    rows = cursor.fetchall()

    # Close the database connection
    conn.close()

    return rows

def create_excel_for_service_provider(service_provider, rows, output_dir):
    # Create an Excel workbook
    workbook = Workbook()
    worksheet = workbook.active

    # Write headers excluding the "id" column
    headers = ["Date (UTC)", "Asset", "Transaction ID", "Party A", "Party B", "Address A", "Address B", "Amount A", "Amount B", "Value in USD (A)", "Value in USD (B)"]
    worksheet.append(headers)

    # Initialize a list to store the maximum column widths
    column_widths = [len(header) for header in headers]

    # Iterate through rows and update maximum column widths
    for row in rows:
        for i, value in enumerate(row[1:]):  # Start from the second element to skip "id"
            value_length = len(str(value))
            if value_length > column_widths[i]:
                column_widths[i] = value_length
        worksheet.append(row[1:])  # Exclude the first element (id) from each row

    # Apply filters to the header row
    worksheet.auto_filter.ref = worksheet.dimensions

    # Set column widths based on the maximum width of values in each column
    for i, width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(i + 1)].width = width + 2

    # Save the Excel file with the service provider's name
    file_name = os.path.join(output_dir, f"{service_provider}.xlsx")
    workbook.save(file_name)

def export_xlsx(db_file, output_dir):
    # Get unique service providers
    service_providers = get_service_providers(db_file)

    # Create the output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Iterate through service providers
    for service_provider in service_providers:
        # Fetch rows for the service provider
        rows = fetch_rows_by_service_provider(db_file, service_provider)

        # Create an Excel file for the service provider
        create_excel_for_service_provider(service_provider, rows, output_dir)

if __name__ == "__main__":
    root = tk.Tk()
    app = GUI(root)

    # Bind the closing event to the on_closing function
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    
    root.mainloop()
